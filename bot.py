"""
Главный файл Telegram бота для управления пользователями Emby
"""

import os
import logging
from datetime import datetime
from typing import List
import asyncio

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    filters
)
from openpyxl import load_workbook

from database import Database
from emby_api import EmbyAPI

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

db = Database()
emby_api = None


def require_admin(func):
    """Декоратор для проверки прав администратора"""
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if not db.is_admin(user_id):
            if update.message:
                await update.message.reply_text("❌ У вас нет прав администратора")
            elif update.callback_query:
                await update.callback_query.answer("❌ У вас нет прав администратора", show_alert=True)
            return
        return await func(update, context)
    return wrapper


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    user_id = update.effective_user.id
    username = update.effective_user.username
    
    if not db.is_admin(user_id):
        await update.message.reply_text(
            "❌ У вас нет доступа к этому боту. Бот доступен только для администраторов."
        )
        return
    
    keyboard = [
        [
            InlineKeyboardButton("📤 Загрузить Excel", callback_data="upload_excel"),
            InlineKeyboardButton("📊 Статистика", callback_data="stats")
        ],
        [
            InlineKeyboardButton("👥 Список пользователей", callback_data="list_users"),
            InlineKeyboardButton("🔍 Проверка входов", callback_data="check_logins")
        ],
        [
            InlineKeyboardButton("👨‍💼 Управление админами", callback_data="manage_admins"),
            InlineKeyboardButton("⚙️ Настройки", callback_data="settings")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        f"👋 Добро пожаловать, {username}!\n\n"
        "🤖 Я бот для управления пользователями Emby\n\n"
        "Основные функции:\n"
        "• Создание пользователей из Excel файлов\n"
        "• Автоматическое удаление через 14 дней после первого входа\n"
        "• Просмотр статистики просмотра\n"
        "• Уведомления администраторам\n\n"
        "Выберите действие:",
        reply_markup=reply_markup
    )


@require_admin
async def add_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавляет администратора"""
    if len(context.args) < 1:
        await update.message.reply_text(
            "⚠️ Использование: /add_admin <telegram_id> [username]\n"
            "Пример: /add_admin 123456789 john"
        )
        return
    
    try:
        telegram_id = int(context.args[0])
        username = context.args[1] if len(context.args) > 1 else None
        
        if db.add_admin(telegram_id, username):
            await update.message.reply_text(f"✅ Администратор {telegram_id} добавлен")
        else:
            await update.message.reply_text(f"⚠️ Администратор {telegram_id} уже существует")
    except ValueError:
        await update.message.reply_text("❌ Неверный формат Telegram ID")


@require_admin
async def remove_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удаляет администратора"""
    if len(context.args) < 1:
        await update.message.reply_text(
            "⚠️ Использование: /remove_admin <telegram_id>\n"
            "Пример: /remove_admin 123456789"
        )
        return
    
    try:
        telegram_id = int(context.args[0])
        
        if db.remove_admin(telegram_id):
            await update.message.reply_text(f"✅ Администратор {telegram_id} удален")
        else:
            await update.message.reply_text(f"⚠️ Администратор {telegram_id} не найден")
    except ValueError:
        await update.message.reply_text("❌ Неверный формат Telegram ID")


@require_admin
async def add_admin_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавляет группу администраторов"""
    if len(context.args) < 1:
        await update.message.reply_text(
            "⚠️ Использование: /add_admin_group <group_id>\n"
            "Пример: /add_admin_group -1001234567890"
        )
        return
    
    try:
        group_id = int(context.args[0])
        
        if db.add_admin_group(group_id):
            await update.message.reply_text(f"✅ Группа {group_id} добавлена")
        else:
            await update.message.reply_text(f"⚠️ Группа {group_id} уже существует")
    except ValueError:
        await update.message.reply_text("❌ Неверный формат ID группы")


@require_admin
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик загруженных документов (Excel файлов)"""
    document = update.message.document
    
    if not (document.file_name.endswith('.xlsx') or document.file_name.endswith('.xls')):
        await update.message.reply_text("❌ Пожалуйста, загрузите Excel файл (.xlsx или .xls)")
        return
    
    await update.message.reply_text("📥 Загружаю файл...")
    
    file = await context.bot.get_file(document.file_id)
    file_path = f"temp_{document.file_name}"
    await file.download_to_drive(file_path)
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        user_col = None
        pass_col = None
        
        for col in range(1, sheet.max_column + 1):
            header = str(sheet.cell(1, col).value).lower() if sheet.cell(1, col).value else ""
            if header == "user":
                user_col = col
            elif header == "pass":
                pass_col = col
        
        if user_col is None or pass_col is None:
            await update.message.reply_text(
                "❌ Не найдены колонки 'user' и 'pass' в Excel файле.\n"
                "Убедитесь, что первая строка содержит заголовки 'user' и 'pass'"
            )
            os.remove(file_path)
            return
        
        created = 0
        errors = 0
        error_messages = []
        
        for row in range(2, sheet.max_row + 1):
            username = sheet.cell(row, user_col).value
            password = sheet.cell(row, pass_col).value
            
            if not username or not password:
                continue
            
            username = str(username).strip()
            password = str(password).strip()
            
            if not username.startswith("user"):
                error_messages.append(f"⚠️ Пропущен {username} (имя не начинается с 'user')")
                errors += 1
                continue
            
            user_data = emby_api.create_user(username, password)
            if user_data:
                emby_user_id = user_data.get('Id')
                db.add_emby_user(username, emby_user_id)
                created += 1
                logger.info(f"✅ Создан пользователь {username}")
            else:
                error_messages.append(f"❌ Ошибка создания {username}")
                errors += 1
        
        report = f"📊 Результаты создания пользователей:\n\n"
        report += f"✅ Создано: {created}\n"
        if errors > 0:
            report += f"❌ Ошибок: {errors}\n\n"
            report += "Детали:\n" + "\n".join(error_messages[:10])
            if len(error_messages) > 10:
                report += f"\n... и еще {len(error_messages) - 10} ошибок"
        
        await update.message.reply_text(report)
        
        os.remove(file_path)
        
    except Exception as e:
        logger.error(f"Ошибка при обработке Excel файла: {e}")
        await update.message.reply_text(f"❌ Ошибка при обработке файла: {str(e)}")
        if os.path.exists(file_path):
            os.remove(file_path)


async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик нажатий на inline кнопки"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    if not db.is_admin(user_id):
        await query.edit_message_text("❌ У вас нет прав администратора")
        return
    
    data = query.data
    
    if data == "upload_excel":
        await query.edit_message_text(
            "📤 Загрузка Excel файла\n\n"
            "Отправьте Excel файл (.xlsx или .xls) со следующей структурой:\n\n"
            "Колонка 'user' - имена пользователей (должны начинаться с 'user')\n"
            "Колонка 'pass' - пароли пользователей\n\n"
            "Первая строка должна содержать заголовки."
        )
    
    elif data == "stats":
        users = db.get_all_users()
        active_users = [u for u in users if not u[3]]
        users_with_login = [u for u in users if u[2] is not None]
        
        text = f"📊 Общая статистика\n\n"
        text += f"👥 Всего пользователей: {len(users)}\n"
        text += f"✅ Активных: {len(active_users)}\n"
        text += f"🚪 Входили: {len(users_with_login)}\n"
        text += f"❌ Удалено: {len(users) - len(active_users)}\n"
        
        await query.edit_message_text(text)
    
    elif data == "list_users":
        users = db.get_all_users()
        
        if not users:
            await query.edit_message_text("📋 Пользователей не найдено")
            return
        
        text = "👥 Список пользователей:\n\n"
        for i, (username, emby_id, first_login, is_deleted) in enumerate(users[:20], 1):
            status = "❌" if is_deleted else "✅"
            login_info = ""
            if first_login:
                login_date = datetime.fromisoformat(first_login) if isinstance(first_login, str) else first_login
                login_info = f" | Вход: {login_date.strftime('%d.%m.%Y')}"
            text += f"{i}. {status} {username}{login_info}\n"
        
        if len(users) > 20:
            text += f"\n... и еще {len(users) - 20} пользователей"
        
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="back_to_menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(text, reply_markup=reply_markup)
    
    elif data == "check_logins":
        await query.edit_message_text("🔍 Проверяю первые входы пользователей...")
        
        users = db.get_all_users()
        checked = 0
        updated = 0
        
        for username, emby_id, first_login, is_deleted in users:
            if is_deleted or first_login:
                continue
            
            login_time = emby_api.check_user_first_login(emby_id)
            if login_time:
                db.update_first_login(emby_id, login_time)
                updated += 1
            
            checked += 1
        
        await query.edit_message_text(
            f"✅ Проверка завершена\n\n"
            f"Проверено пользователей: {checked}\n"
            f"Обновлено записей: {updated}"
        )
    
    elif data == "manage_admins":
        admins = db.get_all_admins()
        groups = db.get_all_admin_groups()
        
        text = "👨‍💼 Управление администраторами\n\n"
        text += f"Администраторов: {len(admins)}\n"
        text += f"Групп: {len(groups)}\n\n"
        text += "Команды:\n"
        text += "/add_admin <id> - добавить админа\n"
        text += "/remove_admin <id> - удалить админа\n"
        text += "/add_admin_group <id> - добавить группу\n"
        
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="back_to_menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(text, reply_markup=reply_markup)
    
    elif data == "settings":
        connection_status = "✅ Подключено" if emby_api.test_connection() else "❌ Нет подключения"
        
        text = f"⚙️ Настройки\n\n"
        text += f"Emby сервер: {connection_status}\n"
        text += f"База данных: ✅ Активна\n"
        text += f"Автоудаление: ✅ Через 14 дней\n"
        
        keyboard = [[InlineKeyboardButton("🔙 Назад", callback_data="back_to_menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(text, reply_markup=reply_markup)
    
    elif data == "back_to_menu":
        keyboard = [
            [
                InlineKeyboardButton("📤 Загрузить Excel", callback_data="upload_excel"),
                InlineKeyboardButton("📊 Статистика", callback_data="stats")
            ],
            [
                InlineKeyboardButton("👥 Список пользователей", callback_data="list_users"),
                InlineKeyboardButton("🔍 Проверка входов", callback_data="check_logins")
            ],
            [
                InlineKeyboardButton("👨‍💼 Управление админами", callback_data="manage_admins"),
                InlineKeyboardButton("⚙️ Настройки", callback_data="settings")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "🤖 Главное меню\n\nВыберите действие:",
            reply_markup=reply_markup
        )


async def check_and_delete_users(context: ContextTypes.DEFAULT_TYPE):
    """
    Фоновая задача: проверяет пользователей и удаляет тех,
    у кого прошло 14 дней с первого входа
    """
    global emby_api
    
    if emby_api is None:
        logger.error("❌ Emby API не инициализирован")
        return
    
    logger.info("🔍 Запуск проверки пользователей для удаления...")
    
    users_to_delete = db.get_users_to_delete(days=14)
    
    if not users_to_delete:
        logger.info("✅ Нет пользователей для удаления")
        return
    
    logger.info(f"📋 Найдено {len(users_to_delete)} пользователей для удаления")
    
    admins = db.get_all_admins()
    admin_groups = db.get_all_admin_groups()
    
    for username, emby_user_id, first_login_at in users_to_delete:
        if emby_api.delete_user(emby_user_id):
            db.mark_user_as_deleted(emby_user_id)
            
            first_login_date = datetime.fromisoformat(first_login_at) if isinstance(first_login_at, str) else first_login_at
            notification = (
                f"🗑 Пользователь удален\n\n"
                f"Имя: {username}\n"
                f"Первый вход: {first_login_date.strftime('%d.%m.%Y %H:%M')}\n"
                f"Причина: Прошло 14 дней с первого входа"
            )
            
            for admin_id in admins:
                try:
                    await context.bot.send_message(
                        chat_id=admin_id,
                        text=notification
                    )
                except Exception as e:
                    logger.error(f"❌ Ошибка отправки уведомления админу {admin_id}: {e}")
            
            for group_id in admin_groups:
                try:
                    await context.bot.send_message(
                        chat_id=group_id,
                        text=notification
                    )
                except Exception as e:
                    logger.error(f"❌ Ошибка отправки уведомления в группу {group_id}: {e}")
            
            logger.info(f"✅ Пользователь {username} удален и уведомления отправлены")


async def check_user_logins(context: ContextTypes.DEFAULT_TYPE):
    """
    Фоновая задача: проверяет первые входы пользователей в Emby
    """
    global emby_api
    
    if emby_api is None:
        logger.error("❌ Emby API не инициализирован")
        return
    
    logger.info("🔍 Проверка первых входов пользователей...")
    
    users = db.get_all_users()
    updated = 0
    
    for username, emby_id, first_login, is_deleted in users:
        if is_deleted or first_login:
            continue
        
        login_time = emby_api.check_user_first_login(emby_id)
        if login_time:
            db.update_first_login(emby_id, login_time)
            updated += 1
            logger.info(f"✅ Обновлен первый вход для {username}")
    
    if updated > 0:
        logger.info(f"✅ Обновлено {updated} записей о первых входах")


def main():
    """Главная функция запуска бота"""
    global emby_api
    
    telegram_token = os.getenv('TELEGRAM_BOT_TOKEN')
    emby_server_url = os.getenv('EMBY_SERVER_URL')
    emby_api_key = os.getenv('EMBY_API_KEY')
    first_admin_id = os.getenv('FIRST_ADMIN_ID')
    
    if not telegram_token:
        logger.error("❌ TELEGRAM_BOT_TOKEN не установлен!")
        return
    
    if not emby_server_url or not emby_api_key:
        logger.error("❌ EMBY_SERVER_URL или EMBY_API_KEY не установлены!")
        return
    
    emby_api = EmbyAPI(emby_server_url, emby_api_key)
    
    if not emby_api.test_connection():
        logger.warning("⚠️ Не удалось подключиться к Emby серверу при запуске. Бот будет работать, но функции Emby недоступны.")
        logger.warning("⚠️ Проверьте, что Emby сервер доступен из облака Replit, или используйте публичный URL.")
    else:
        logger.info("✅ Успешное подключение к Emby серверу!")
    
    if first_admin_id:
        try:
            db.add_admin(int(first_admin_id))
            logger.info(f"✅ Первый администратор {first_admin_id} добавлен")
        except ValueError:
            logger.error("❌ Неверный формат FIRST_ADMIN_ID")
    
    application = Application.builder().token(telegram_token).build()
    
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("add_admin", add_admin))
    application.add_handler(CommandHandler("remove_admin", remove_admin))
    application.add_handler(CommandHandler("add_admin_group", add_admin_group))
    
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    application.add_handler(CallbackQueryHandler(button_callback))
    
    application.job_queue.run_repeating(check_user_logins, interval=3600, first=10)
    
    application.job_queue.run_repeating(check_and_delete_users, interval=21600, first=60)
    
    logger.info("🤖 Бот запущен!")
    
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()
