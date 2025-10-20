"""
–ì–ª–∞–≤–Ω—ã–π —Ñ–∞–π–ª Telegram –±–æ—Ç–∞ –¥–ª—è —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è–º–∏ Emby
"""

import os
import logging
from datetime import datetime
from typing import List
import asyncio

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    filters
)
from openpyxl import load_workbook

from database import Database
from emby_api import EmbyAPI

# –ù–∞—Å—Ç—Ä–æ–π–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è
db = Database()
emby_api = None


def require_admin(func):
    """–î–µ–∫–æ—Ä–∞—Ç–æ—Ä –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ –ø—Ä–∞–≤ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞"""
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if not db.is_admin(user_id):
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ç–∏–ø update –∏ –æ—Ç–≤–µ—á–∞–µ–º —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤–µ–Ω–Ω–æ
            if update.message:
                await update.message.reply_text("‚ùå –£ –≤–∞—Å –Ω–µ—Ç –ø—Ä–∞–≤ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞")
            elif update.callback_query:
                await update.callback_query.answer("‚ùå –£ –≤–∞—Å –Ω–µ—Ç –ø—Ä–∞–≤ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞", show_alert=True)
            return
        return await func(update, context)
    return wrapper


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∫–æ–º–∞–Ω–¥—ã /start"""
    user_id = update.effective_user.id
    username = update.effective_user.username
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —è–≤–ª—è–µ—Ç—Å—è –ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–º
    if not db.is_admin(user_id):
        await update.message.reply_text(
            "‚ùå –£ –≤–∞—Å –Ω–µ—Ç –¥–æ—Å—Ç—É–ø–∞ –∫ —ç—Ç–æ–º—É –±–æ—Ç—É. –ë–æ—Ç –¥–æ—Å—Ç—É–ø–µ–Ω —Ç–æ–ª—å–∫–æ –¥–ª—è –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–≤."
        )
        return
    
    keyboard = [
        [
            InlineKeyboardButton("üì§ –ó–∞–≥—Ä—É–∑–∏—Ç—å Excel", callback_data="upload_excel"),
            InlineKeyboardButton("üìä –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞", callback_data="stats")
        ],
        [
            InlineKeyboardButton("üë• –°–ø–∏—Å–æ–∫ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π", callback_data="list_users"),
            InlineKeyboardButton("üîç –ü—Ä–æ–≤–µ—Ä–∫–∞ –≤—Ö–æ–¥–æ–≤", callback_data="check_logins")
        ],
        [
            InlineKeyboardButton("üë®‚Äçüíº –£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∞–¥–º–∏–Ω–∞–º–∏", callback_data="manage_admins"),
            InlineKeyboardButton("‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏", callback_data="settings")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        f"üëã –î–æ–±—Ä–æ –ø–æ–∂–∞–ª–æ–≤–∞—Ç—å, {username}!\n\n"
        "ü§ñ –Ø –±–æ—Ç –¥–ª—è —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è–º–∏ Emby\n\n"
        "–û—Å–Ω–æ–≤–Ω—ã–µ —Ñ—É–Ω–∫—Ü–∏–∏:\n"
        "‚Ä¢ –°–æ–∑–¥–∞–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏–∑ Excel —Ñ–∞–π–ª–æ–≤\n"
        "‚Ä¢ –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–æ–µ —É–¥–∞–ª–µ–Ω–∏–µ —á–µ—Ä–µ–∑ 14 –¥–Ω–µ–π –ø–æ—Å–ª–µ –ø–µ—Ä–≤–æ–≥–æ –≤—Ö–æ–¥–∞\n"
        "‚Ä¢ –ü—Ä–æ—Å–º–æ—Ç—Ä —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –ø—Ä–æ—Å–º–æ—Ç—Ä–∞\n"
        "‚Ä¢ –£–≤–µ–¥–æ–º–ª–µ–Ω–∏—è –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º\n\n"
        "–í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:",
        reply_markup=reply_markup
    )


@require_admin
async def add_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """–î–æ–±–∞–≤–ª—è–µ—Ç –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞"""
    if len(context.args) < 1:
        await update.message.reply_text(
            "‚ö†Ô∏è –ò—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–µ: /add_admin <telegram_id> [username]\n"
            "–ü—Ä–∏–º–µ—Ä: /add_admin 123456789 john"
        )
        return
    
    try:
        telegram_id = int(context.args[0])
        username = context.args[1] if len(context.args) > 1 else None
        
        if db.add_admin(telegram_id, username):
            await update.message.reply_text(f"‚úÖ –ê–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä {telegram_id} –¥–æ–±–∞–≤–ª–µ–Ω")
        else:
            await update.message.reply_text(f"‚ö†Ô∏è –ê–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä {telegram_id} —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç")
    except ValueError:
        await update.message.reply_text("‚ùå –ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç Telegram ID")


@require_admin
async def remove_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """–£–¥–∞–ª—è–µ—Ç –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞"""
    if len(context.args) < 1:
        await update.message.reply_text(
            "‚ö†Ô∏è –ò—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–µ: /remove_admin <telegram_id>\n"
            "–ü—Ä–∏–º–µ—Ä: /remove_admin 123456789"
        )
        return
    
    try:
        telegram_id = int(context.args[0])
        
        if db.remove_admin(telegram_id):
            await update.message.reply_text(f"‚úÖ –ê–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä {telegram_id} —É–¥–∞–ª–µ–Ω")
        else:
            await update.message.reply_text(f"‚ö†Ô∏è –ê–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä {telegram_id} –Ω–µ –Ω–∞–π–¥–µ–Ω")
    except ValueError:
        await update.message.reply_text("‚ùå –ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç Telegram ID")


@require_admin
async def add_admin_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """–î–æ–±–∞–≤–ª—è–µ—Ç –≥—Ä—É–ø–ø—É –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–≤"""
    if len(context.args) < 1:
        await update.message.reply_text(
            "‚ö†Ô∏è –ò—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–µ: /add_admin_group <group_id>\n"
            "–ü—Ä–∏–º–µ—Ä: /add_admin_group -1001234567890"
        )
        return
    
    try:
        group_id = int(context.args[0])
        
        if db.add_admin_group(group_id):
            await update.message.reply_text(f"‚úÖ –ì—Ä—É–ø–ø–∞ {group_id} –¥–æ–±–∞–≤–ª–µ–Ω–∞")
        else:
            await update.message.reply_text(f"‚ö†Ô∏è –ì—Ä—É–ø–ø–∞ {group_id} —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç")
    except ValueError:
        await update.message.reply_text("‚ùå –ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç ID –≥—Ä—É–ø–ø—ã")


@require_admin
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤ (Excel —Ñ–∞–π–ª–æ–≤)"""
    document = update.message.document
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ —ç—Ç–æ Excel —Ñ–∞–π–ª
    if not (document.file_name.endswith('.xlsx') or document.file_name.endswith('.xls')):
        await update.message.reply_text("‚ùå –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –∑–∞–≥—Ä—É–∑–∏—Ç–µ Excel —Ñ–∞–π–ª (.xlsx –∏–ª–∏ .xls)")
        return
    
    await update.message.reply_text("üì• –ó–∞–≥—Ä—É–∂–∞—é —Ñ–∞–π–ª...")
    
    # –°–∫–∞—á–∏–≤–∞–µ–º —Ñ–∞–π–ª
    file = await context.bot.get_file(document.file_id)
    file_path = f"temp_{document.file_name}"
    await file.download_to_drive(file_path)
    
    try:
        # –ß–∏—Ç–∞–µ–º Excel —Ñ–∞–π–ª
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # –ò—â–µ–º –∫–æ–ª–æ–Ω–∫–∏ "user" –∏ "pass"
        user_col = None
        pass_col = None
        
        for col in range(1, sheet.max_column + 1):
            header = str(sheet.cell(1, col).value).lower() if sheet.cell(1, col).value else ""
            if header == "user":
                user_col = col
            elif header == "pass":
                pass_col = col
        
        if user_col is None or pass_col is None:
            await update.message.reply_text(
                "‚ùå –ù–µ –Ω–∞–π–¥–µ–Ω—ã –∫–æ–ª–æ–Ω–∫–∏ 'user' –∏ 'pass' –≤ Excel —Ñ–∞–π–ª–µ.\n"
                "–£–±–µ–¥–∏—Ç–µ—Å—å, —á—Ç–æ –ø–µ—Ä–≤–∞—è —Å—Ç—Ä–æ–∫–∞ —Å–æ–¥–µ—Ä–∂–∏—Ç –∑–∞–≥–æ–ª–æ–≤–∫–∏ 'user' –∏ 'pass'"
            )
            os.remove(file_path)
            return
        
        # –°–æ–∑–¥–∞–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
        created = 0
        errors = 0
        error_messages = []
        
        for row in range(2, sheet.max_row + 1):
            username = sheet.cell(row, user_col).value
            password = sheet.cell(row, pass_col).value
            
            if not username or not password:
                continue
            
            username = str(username).strip()
            password = str(password).strip()
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –∏–º—è –Ω–∞—á–∏–Ω–∞–µ—Ç—Å—è —Å "user"
            if not username.startswith("user"):
                error_messages.append(f"‚ö†Ô∏è –ü—Ä–æ–ø—É—â–µ–Ω {username} (–∏–º—è –Ω–µ –Ω–∞—á–∏–Ω–∞–µ—Ç—Å—è —Å 'user')")
                errors += 1
                continue
            
            # –°–æ–∑–¥–∞–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ Emby
            user_data = emby_api.create_user(username, password)
            if user_data:
                emby_user_id = user_data.get('Id')
                db.add_emby_user(username, emby_user_id)
                created += 1
                logger.info(f"‚úÖ –°–æ–∑–¥–∞–Ω –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å {username}")
            else:
                error_messages.append(f"‚ùå –û—à–∏–±–∫–∞ —Å–æ–∑–¥–∞–Ω–∏—è {username}")
                errors += 1
        
        # –§–æ—Ä–º–∏—Ä—É–µ–º –æ—Ç—á–µ—Ç
        report = f"üìä –†–µ–∑—É–ª—å—Ç–∞—Ç—ã —Å–æ–∑–¥–∞–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π:\n\n"
        report += f"‚úÖ –°–æ–∑–¥–∞–Ω–æ: {created}\n"
        if errors > 0:
            report += f"‚ùå –û—à–∏–±–æ–∫: {errors}\n\n"
            report += "–î–µ—Ç–∞–ª–∏:\n" + "\n".join(error_messages[:10])
            if len(error_messages) > 10:
                report += f"\n... –∏ –µ—â–µ {len(error_messages) - 10} –æ—à–∏–±–æ–∫"
        
        await update.message.reply_text(report)
        
        # –£–¥–∞–ª—è–µ–º –≤—Ä–µ–º–µ–Ω–Ω—ã–π —Ñ–∞–π–ª
        os.remove(file_path)
        
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ Excel —Ñ–∞–π–ª–∞: {e}")
        await update.message.reply_text(f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ —Ñ–∞–π–ª–∞: {str(e)}")
        if os.path.exists(file_path):
            os.remove(file_path)


async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –Ω–∞–∂–∞—Ç–∏–π –Ω–∞ inline –∫–Ω–æ–ø–∫–∏"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    if not db.is_admin(user_id):
        await query.edit_message_text("‚ùå –£ –≤–∞—Å –Ω–µ—Ç –ø—Ä–∞–≤ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞")
        return
    
    data = query.data
    
    if data == "upload_excel":
        await query.edit_message_text(
            "üì§ –ó–∞–≥—Ä—É–∑–∫–∞ Excel —Ñ–∞–π–ª–∞\n\n"
            "–û—Ç–ø—Ä–∞–≤—å—Ç–µ Excel —Ñ–∞–π–ª (.xlsx –∏–ª–∏ .xls) —Å–æ —Å–ª–µ–¥—É—é—â–µ–π —Å—Ç—Ä—É–∫—Ç—É—Ä–æ–π:\n\n"
            "–ö–æ–ª–æ–Ω–∫–∞ 'user' - –∏–º–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π (–¥–æ–ª–∂–Ω—ã –Ω–∞—á–∏–Ω–∞—Ç—å—Å—è —Å 'user')\n"
            "–ö–æ–ª–æ–Ω–∫–∞ 'pass' - –ø–∞—Ä–æ–ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π\n\n"
            "–ü–µ—Ä–≤–∞—è —Å—Ç—Ä–æ–∫–∞ –¥–æ–ª–∂–Ω–∞ —Å–æ–¥–µ—Ä–∂–∞—Ç—å –∑–∞–≥–æ–ª–æ–≤–∫–∏."
        )
    
    elif data == "stats":
        users = db.get_all_users()
        active_users = [u for u in users if not u[3]]  # is_deleted == False
        users_with_login = [u for u in users if u[2] is not None]  # first_login_at is not None
        
        text = f"üìä –û–±—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞\n\n"
        text += f"üë• –í—Å–µ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π: {len(users)}\n"
        text += f"‚úÖ –ê–∫—Ç–∏–≤–Ω—ã—Ö: {len(active_users)}\n"
        text += f"üö™ –í—Ö–æ–¥–∏–ª–∏: {len(users_with_login)}\n"
        text += f"‚ùå –£–¥–∞–ª–µ–Ω–æ: {len(users) - len(active_users)}\n"
        
        await query.edit_message_text(text)
    
    elif data == "list_users":
        users = db.get_all_users()
        
        if not users:
            await query.edit_message_text("üìã –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –Ω–µ –Ω–∞–π–¥–µ–Ω–æ")
            return
        
        text = "üë• –°–ø–∏—Å–æ–∫ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π:\n\n"
        for i, (username, emby_id, first_login, is_deleted) in enumerate(users[:20], 1):
            status = "‚ùå" if is_deleted else "‚úÖ"
            login_info = ""
            if first_login:
                login_date = datetime.fromisoformat(first_login) if isinstance(first_login, str) else first_login
                login_info = f" | –í—Ö–æ–¥: {login_date.strftime('%d.%m.%Y')}"
            text += f"{i}. {status} {username}{login_info}\n"
        
        if len(users) > 20:
            text += f"\n... –∏ –µ—â–µ {len(users) - 20} –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"
        
        keyboard = [[InlineKeyboardButton("üîô –ù–∞–∑–∞–¥", callback_data="back_to_menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(text, reply_markup=reply_markup)
    
    elif data == "check_logins":
        await query.edit_message_text("üîç –ü—Ä–æ–≤–µ—Ä—è—é –ø–µ—Ä–≤—ã–µ –≤—Ö–æ–¥—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π...")
        
        users = db.get_all_users()
        checked = 0
        updated = 0
        
        for username, emby_id, first_login, is_deleted in users:
            if is_deleted or first_login:
                continue
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–µ—Ä–≤—ã–π –≤—Ö–æ–¥ —á–µ—Ä–µ–∑ Emby API
            login_time = emby_api.check_user_first_login(emby_id)
            if login_time:
                db.update_first_login(emby_id, login_time)
                updated += 1
            
            checked += 1
        
        await query.edit_message_text(
            f"‚úÖ –ü—Ä–æ–≤–µ—Ä–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞\n\n"
            f"–ü—Ä–æ–≤–µ—Ä–µ–Ω–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π: {checked}\n"
            f"–û–±–Ω–æ–≤–ª–µ–Ω–æ –∑–∞–ø–∏—Å–µ–π: {updated}"
        )
    
    elif data == "manage_admins":
        admins = db.get_all_admins()
        groups = db.get_all_admin_groups()
        
        text = "üë®‚Äçüíº –£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º–∏\n\n"
        text += f"–ê–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–≤: {len(admins)}\n"
        text += f"–ì—Ä—É–ø–ø: {len(groups)}\n\n"
        text += "–ö–æ–º–∞–Ω–¥—ã:\n"
        text += "/add_admin <id> - –¥–æ–±–∞–≤–∏—Ç—å –∞–¥–º–∏–Ω–∞\n"
        text += "/remove_admin <id> - —É–¥–∞–ª–∏—Ç—å –∞–¥–º–∏–Ω–∞\n"
        text += "/add_admin_group <id> - –¥–æ–±–∞–≤–∏—Ç—å –≥—Ä—É–ø–ø—É\n"
        
        keyboard = [[InlineKeyboardButton("üîô –ù–∞–∑–∞–¥", callback_data="back_to_menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(text, reply_markup=reply_markup)
    
    elif data == "settings":
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–æ–¥–∫–ª—é—á–µ–Ω–∏–µ –∫ Emby
        connection_status = "‚úÖ –ü–æ–¥–∫–ª—é—á–µ–Ω–æ" if emby_api.test_connection() else "‚ùå –ù–µ—Ç –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è"
        
        text = f"‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏\n\n"
        text += f"Emby —Å–µ—Ä–≤–µ—Ä: {connection_status}\n"
        text += f"–ë–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö: ‚úÖ –ê–∫—Ç–∏–≤–Ω–∞\n"
        text += f"–ê–≤—Ç–æ—É–¥–∞–ª–µ–Ω–∏–µ: ‚úÖ –ß–µ—Ä–µ–∑ 14 –¥–Ω–µ–π\n"
        
        keyboard = [[InlineKeyboardButton("üîô –ù–∞–∑–∞–¥", callback_data="back_to_menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(text, reply_markup=reply_markup)
    
    elif data == "back_to_menu":
        keyboard = [
            [
                InlineKeyboardButton("üì§ –ó–∞–≥—Ä—É–∑–∏—Ç—å Excel", callback_data="upload_excel"),
                InlineKeyboardButton("üìä –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞", callback_data="stats")
            ],
            [
                InlineKeyboardButton("üë• –°–ø–∏—Å–æ–∫ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π", callback_data="list_users"),
                InlineKeyboardButton("üîç –ü—Ä–æ–≤–µ—Ä–∫–∞ –≤—Ö–æ–¥–æ–≤", callback_data="check_logins")
            ],
            [
                InlineKeyboardButton("üë®‚Äçüíº –£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∞–¥–º–∏–Ω–∞–º–∏", callback_data="manage_admins"),
                InlineKeyboardButton("‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏", callback_data="settings")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "ü§ñ –ì–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é\n\n–í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:",
            reply_markup=reply_markup
        )


async def check_and_delete_users(context: ContextTypes.DEFAULT_TYPE):
    """
    –§–æ–Ω–æ–≤–∞—è –∑–∞–¥–∞—á–∞: –ø—Ä–æ–≤–µ—Ä—è–µ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏ —É–¥–∞–ª—è–µ—Ç —Ç–µ—Ö,
    —É –∫–æ–≥–æ –ø—Ä–æ—à–ª–æ 14 –¥–Ω–µ–π —Å –ø–µ—Ä–≤–æ–≥–æ –≤—Ö–æ–¥–∞
    """
    global emby_api
    
    if emby_api is None:
        logger.error("‚ùå Emby API –Ω–µ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω")
        return
    
    logger.info("üîç –ó–∞–ø—É—Å–∫ –ø—Ä–æ–≤–µ—Ä–∫–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è...")
    
    # –ü–æ–ª—É—á–∞–µ–º —Å–ø–∏—Å–æ–∫ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è
    users_to_delete = db.get_users_to_delete(days=14)
    
    if not users_to_delete:
        logger.info("‚úÖ –ù–µ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è")
        return
    
    logger.info(f"üìã –ù–∞–π–¥–µ–Ω–æ {len(users_to_delete)} –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è")
    
    # –ü–æ–ª—É—á–∞–µ–º —Å–ø–∏—Å–æ–∫ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–≤ –¥–ª—è —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–π
    admins = db.get_all_admins()
    admin_groups = db.get_all_admin_groups()
    
    for username, emby_user_id, first_login_at in users_to_delete:
        # –£–¥–∞–ª—è–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –∏–∑ Emby
        if emby_api.delete_user(emby_user_id):
            # –û—Ç–º–µ—á–∞–µ–º –∫–∞–∫ —É–¥–∞–ª–µ–Ω–Ω–æ–≥–æ –≤ –ë–î
            db.mark_user_as_deleted(emby_user_id)
            
            # –§–æ—Ä–º–∏—Ä—É–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è
            first_login_date = datetime.fromisoformat(first_login_at) if isinstance(first_login_at, str) else first_login_at
            notification = (
                f"üóë –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —É–¥–∞–ª–µ–Ω\n\n"
                f"–ò–º—è: {username}\n"
                f"–ü–µ—Ä–≤—ã–π –≤—Ö–æ–¥: {first_login_date.strftime('%d.%m.%Y %H:%M')}\n"
                f"–ü—Ä–∏—á–∏–Ω–∞: –ü—Ä–æ—à–ª–æ 14 –¥–Ω–µ–π —Å –ø–µ—Ä–≤–æ–≥–æ –≤—Ö–æ–¥–∞"
            )
            
            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º –≤ –õ–°
            for admin_id in admins:
                try:
                    await context.bot.send_message(
                        chat_id=admin_id,
                        text=notification
                    )
                except Exception as e:
                    logger.error(f"‚ùå –û—à–∏–±–∫–∞ –æ—Ç–ø—Ä–∞–≤–∫–∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è –∞–¥–º–∏–Ω—É {admin_id}: {e}")
            
            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è –≤ –≥—Ä—É–ø–ø—ã –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–≤
            for group_id in admin_groups:
                try:
                    await context.bot.send_message(
                        chat_id=group_id,
                        text=notification
                    )
                except Exception as e:
                    logger.error(f"‚ùå –û—à–∏–±–∫–∞ –æ—Ç–ø—Ä–∞–≤–∫–∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è –≤ –≥—Ä—É–ø–ø—É {group_id}: {e}")
            
            logger.info(f"‚úÖ –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å {username} —É–¥–∞–ª–µ–Ω –∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω—ã")


async def check_user_logins(context: ContextTypes.DEFAULT_TYPE):
    """
    –§–æ–Ω–æ–≤–∞—è –∑–∞–¥–∞—á–∞: –ø—Ä–æ–≤–µ—Ä—è–µ—Ç –ø–µ—Ä–≤—ã–µ –≤—Ö–æ–¥—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –≤ Emby
    """
    global emby_api
    
    if emby_api is None:
        logger.error("‚ùå Emby API –Ω–µ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω")
        return
    
    logger.info("üîç –ü—Ä–æ–≤–µ—Ä–∫–∞ –ø–µ—Ä–≤—ã—Ö –≤—Ö–æ–¥–æ–≤ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π...")
    
    users = db.get_all_users()
    updated = 0
    
    for username, emby_id, first_login, is_deleted in users:
        # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º —É–¥–∞–ª–µ–Ω–Ω—ã—Ö –∏ —Ç–µ—Ö, —É –∫–æ–≥–æ —É–∂–µ –∑–∞–ø–∏—Å–∞–Ω –ø–µ—Ä–≤—ã–π –≤—Ö–æ–¥
        if is_deleted or first_login:
            continue
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–µ—Ä–≤—ã–π –≤—Ö–æ–¥ —á–µ—Ä–µ–∑ Emby API
        login_time = emby_api.check_user_first_login(emby_id)
        if login_time:
            db.update_first_login(emby_id, login_time)
            updated += 1
            logger.info(f"‚úÖ –û–±–Ω–æ–≤–ª–µ–Ω –ø–µ—Ä–≤—ã–π –≤—Ö–æ–¥ –¥–ª—è {username}")
    
    if updated > 0:
        logger.info(f"‚úÖ –û–±–Ω–æ–≤–ª–µ–Ω–æ {updated} –∑–∞–ø–∏—Å–µ–π –æ –ø–µ—Ä–≤—ã—Ö –≤—Ö–æ–¥–∞—Ö")


def main():
    """–ì–ª–∞–≤–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è –∑–∞–ø—É—Å–∫–∞ –±–æ—Ç–∞"""
    global emby_api
    
    # –ü–æ–ª—É—á–∞–µ–º –ø–µ—Ä–µ–º–µ–Ω–Ω—ã–µ –æ–∫—Ä—É–∂–µ–Ω–∏—è
    telegram_token = os.getenv('TELEGRAM_BOT_TOKEN')
    emby_server_url = os.getenv('EMBY_SERVER_URL')
    emby_api_key = os.getenv('EMBY_API_KEY')
    first_admin_id = os.getenv('FIRST_ADMIN_ID')
    
    if not telegram_token:
        logger.error("‚ùå TELEGRAM_BOT_TOKEN –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω!")
        return
    
    if not emby_server_url or not emby_api_key:
        logger.error("‚ùå EMBY_SERVER_URL –∏–ª–∏ EMBY_API_KEY –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω—ã!")
        return
    
    # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ–º Emby API
    emby_api = EmbyAPI(emby_server_url, emby_api_key)
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–æ–¥–∫–ª—é—á–µ–Ω–∏–µ –∫ Emby
    if not emby_api.test_connection():
        logger.error("‚ùå –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥–∫–ª—é—á–∏—Ç—å—Å—è –∫ Emby —Å–µ—Ä–≤–µ—Ä—É!")
        return
    
    # –î–æ–±–∞–≤–ª—è–µ–º –ø–µ—Ä–≤–æ–≥–æ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞, –µ—Å–ª–∏ —É–∫–∞–∑–∞–Ω
    if first_admin_id:
        try:
            db.add_admin(int(first_admin_id))
            logger.info(f"‚úÖ –ü–µ—Ä–≤—ã–π –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä {first_admin_id} –¥–æ–±–∞–≤–ª–µ–Ω")
        except ValueError:
            logger.error("‚ùå –ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç FIRST_ADMIN_ID")
    
    # –°–æ–∑–¥–∞–µ–º –ø—Ä–∏–ª–æ–∂–µ–Ω–∏–µ
    application = Application.builder().token(telegram_token).build()
    
    # –î–æ–±–∞–≤–ª—è–µ–º –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–∏ –∫–æ–º–∞–Ω–¥
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("add_admin", add_admin))
    application.add_handler(CommandHandler("remove_admin", remove_admin))
    application.add_handler(CommandHandler("add_admin_group", add_admin_group))
    
    # –û–±—Ä–∞–±–æ—Ç—á–∏–∫ –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤ (Excel —Ñ–∞–π–ª–æ–≤)
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    # –û–±—Ä–∞–±–æ—Ç—á–∏–∫ inline –∫–Ω–æ–ø–æ–∫
    application.add_handler(CallbackQueryHandler(button_callback))
    
    # –ù–∞—Å—Ç—Ä–∞–∏–≤–∞–µ–º —Ñ–æ–Ω–æ–≤—ã–µ –∑–∞–¥–∞—á–∏
    # –ü—Ä–æ–≤–µ—Ä–∫–∞ –ø–µ—Ä–≤—ã—Ö –≤—Ö–æ–¥–æ–≤ –∫–∞–∂–¥—ã–π —á–∞—Å
    application.job_queue.run_repeating(check_user_logins, interval=3600, first=10)
    
    # –ü—Ä–æ–≤–µ—Ä–∫–∞ –∏ —É–¥–∞–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∫–∞–∂–¥—ã–µ 6 —á–∞—Å–æ–≤
    application.job_queue.run_repeating(check_and_delete_users, interval=21600, first=60)
    
    logger.info("ü§ñ –ë–æ—Ç –∑–∞–ø—É—â–µ–Ω!")
    
    # –ó–∞–ø—É—Å–∫–∞–µ–º –±–æ—Ç–∞
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()
